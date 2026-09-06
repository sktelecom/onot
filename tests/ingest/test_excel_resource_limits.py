# SPDX-FileCopyrightText: Kakao Corp. and SK telecom Co., Ltd.
# SPDX-License-Identifier: Apache-2.0

"""ExcelAdapter resource-exhaustion guards.

An .xlsx is a ZIP, so a tiny upload can inflate to an unbounded worksheet (a repeated-row
payload compresses ~300x) and drive peak memory up while the request still returns 200.
These tests pin the decompression-bomb rejection, the row-scan cap, and the dedup-during-scan
that keeps duplicate rows from ever materializing as distinct Package objects.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from onot.domain.errors import ParseError
from onot.ingest import excel
from onot.ingest.excel import parse_excel

FIXTURE = Path(__file__).parent.parent / "fixtures" / "sbom" / "example.spdx.xlsx"


def _build(path, package_rows, *, with_extracted=False):
    wb = openpyxl.Workbook()
    di = wb.active
    di.title = "Document Info"
    di.append(["h"] * 6)
    di.append(["2.3.0", "SPDX-2.3", "CC0-1.0", "id", "3.17", "MyProduct"])
    pi = wb.create_sheet("Package Info")
    pi.append(["Package Name"])
    for row in package_rows:
        pi.append(row)
    if with_extracted:
        wb.create_sheet("Extracted License Info").append(["Identifier"])
    wb.save(path)


def test_normal_fixture_passes_guard():
    # The real SBOM fixture (overall ratio ~3.3x) must sail through untouched.
    doc = parse_excel(FIXTURE)
    assert len(doc.packages) >= 1


def test_compression_bomb_rejected(tmp_path):
    # A worksheet member that inflates far past the ratio cap is rejected from the ZIP
    # central directory, before openpyxl reads any payload.
    path = tmp_path / "bomb.xlsx"
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as z:
        z.writestr("xl/worksheets/sheet1.xml", b"<row>same</row>" * 2_000_000)
    with pytest.raises(ParseError, match="compression ratio|decompression bomb"):
        parse_excel(path)


def test_uncompressed_size_cap(tmp_path, monkeypatch):
    # Even at a normal ratio, total inflated size beyond the cap is rejected.
    monkeypatch.setattr(excel, "_MAX_TOTAL_UNCOMPRESSED_BYTES", 4096)
    monkeypatch.setattr(excel, "_MAX_COMPRESSION_RATIO", 10_000.0)
    path = tmp_path / "big.xlsx"
    _build(path, [["pkg", "", "1.0"]])
    with pytest.raises(ParseError, match="expands beyond the allowed size"):
        parse_excel(path)


def test_too_many_entries_rejected(tmp_path, monkeypatch):
    monkeypatch.setattr(excel, "_MAX_ZIP_ENTRIES", 3)
    path = tmp_path / "many.xlsx"
    with zipfile.ZipFile(path, "w") as z:
        for i in range(5):
            z.writestr(f"m{i}.xml", "x")
    with pytest.raises(ParseError, match="too many internal entries"):
        parse_excel(path)


def test_row_cap_rejected(tmp_path, monkeypatch):
    monkeypatch.setattr(excel, "_MAX_PACKAGE_ROWS", 3)
    path = tmp_path / "rows.xlsx"
    _build(path, [[f"pkg-{i}", "", str(i)] for i in range(5)])
    with pytest.raises(ParseError, match="too many rows"):
        parse_excel(path)


def test_duplicate_rows_collapsed(tmp_path):
    # Repeated (name, version) rows must not each become a Package; only the first survives.
    path = tmp_path / "dupes.xlsx"
    _build(path, [["dup", "", "1.0"]] * 50 + [["other", "", "2.0"]])
    doc = parse_excel(path)
    assert [(p.name, p.version) for p in doc.packages] == [("dup", "1.0"), ("other", "2.0")]


def test_legacy_non_zip_left_to_openpyxl(tmp_path):
    # A non-ZIP input passes the guard (returns) and is rejected by openpyxl as a ParseError.
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"\xd0\xcf\x11\xe0not-a-real-workbook")
    with pytest.raises(ParseError):
        parse_excel(path)
