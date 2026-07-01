# SPDX-FileCopyrightText: Kakao Corp. and SK telecom Co., Ltd.
# SPDX-License-Identifier: Apache-2.0

"""ExcelAdapter robustness - works without IndexError on short rows/blank rows/missing sheets (L3 #1 regression)."""

from __future__ import annotations

import zipfile

import openpyxl
import pytest

from onot.domain.errors import ParseError
from onot.ingest.excel import parse_excel


def _build(path, package_rows, *, with_extracted=True):
    wb = openpyxl.Workbook()
    di = wb.active
    di.title = "Document Info"
    di.append(["h"] * 6)  # header (>=6 columns; Document Name is index 5)
    di.append(["2.3.0", "SPDX-2.3", "CC0-1.0", "id", "3.17", "MyProduct"])
    pi = wb.create_sheet("Package Info")
    pi.append(["Package Name"])  # header
    for row in package_rows:
        pi.append(row)
    if with_extracted:
        wb.create_sheet("Extracted License Info").append(["Identifier"])
    wb.save(path)


def test_short_rows_do_not_crash(tmp_path):
    path = tmp_path / "mini.xlsx"
    # A 1-cell row (name only) and a 3-cell row (up to version) - both under 16 cells, so
    # fixed-index (12/13/16) access raises IndexError without the _cell guard.
    _build(path, [["foo"], ["bar", "SPDXRef-bar", "1.0"]])
    doc = parse_excel(path)
    assert doc.name == "MyProduct"
    assert [pk.name for pk in doc.packages] == ["foo", "bar"]
    foo = doc.packages[0]
    assert foo.version == ""
    assert foo.copyright is None
    assert foo.effective_expression is None
    assert doc.packages[1].version == "1.0"  # version is column index 2


def test_missing_extracted_sheet(tmp_path):
    path = tmp_path / "noext.xlsx"
    _build(path, [["foo"]], with_extracted=False)
    doc = parse_excel(path)
    assert doc.license_refs == ()


def test_blank_rows_skipped(tmp_path):
    path = tmp_path / "blank.xlsx"
    _build(path, [[None], ["", None], ["keep"]])
    doc = parse_excel(path)
    assert [pk.name for pk in doc.packages] == ["keep"]


def test_non_excel_zip_raises_parse_error(tmp_path):
    # O-5: a zip that is not an xlsx (.zip/.docx/.jar) must be a 400-class ParseError, not a 500.
    path = tmp_path / "fake.zip"
    with zipfile.ZipFile(path, "w") as z:
        z.writestr("a.txt", "hi")
    with pytest.raises(ParseError):
        parse_excel(path)


def test_missing_required_sheet_raises_parse_error(tmp_path):
    # O-6: a valid xlsx without the required sheets must be a ParseError, not a KeyError/500.
    path = tmp_path / "nosheets.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Random"
    wb.save(path)
    with pytest.raises(ParseError, match="required sheet"):
        parse_excel(path)
