# SPDX-FileCopyrightText: Kakao Corp. and SK telecom Co., Ltd.
# SPDX-License-Identifier: Apache-2.0

"""ExcelAdapter — 1.x onot/SPDX spreadsheet template -> NoticeDocument.

Reads the Document Info, Package Info, and Extracted License Info sheets and maps them into
a NoticeDocument. Integrated into the registry as ExcelAdapter. Column indices follow the
standard SPDX spreadsheet schema.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl

from onot.domain.errors import ParseError
from onot.domain.models import Copyright, LicenseExpression, LicenseRef, NoticeDocument, Package
from onot.ingest.base import IngestResult

_REQUIRED_SHEETS = ("Document Info", "Package Info")

# Resource-exhaustion guards. An .xlsx is a ZIP, so a tiny upload can inflate to an
# unbounded worksheet (a repeated-row payload compresses ~300x). The upload byte cap
# (routes.MAX_UPLOAD_BYTES) only bounds the *compressed* size, so these limits bound the
# decompressed cost instead: reject the file from the ZIP central directory before openpyxl
# reads it, and stop the row scan before it builds an unbounded number of Package objects.
_MAX_ZIP_ENTRIES = 1024
_MAX_TOTAL_UNCOMPRESSED_BYTES = 128 * 1024 * 1024  # 128 MiB across all members
_MAX_COMPRESSION_RATIO = 100.0  # overall uncompressed/compressed; normal SBOM xlsx is ~3-6x
_MAX_PACKAGE_ROWS = 100_000  # non-empty Package Info rows scanned before aborting

# Standard SPDX spreadsheet column indices (0-based)
_DOC_NAME_COL = 5
_PKG_COLS = {
    "name": 0,
    "version": 2,
    "download": 7,
    "declared": 12,
    "concluded": 13,
    "copyright": 16,
}
_EXT_COLS = {"id": 0, "text": 1, "name": 2}


class ExcelAdapter:
    format_id = "excel"

    def sniff(self, path: Path, head: bytes) -> float:
        if path.name.lower().endswith((".xlsx", ".xls")):
            return 0.8
        if head.startswith(b"PK\x03\x04"):  # xlsx = zip
            return 0.4
        return 0.0

    def parse(self, path: Path) -> IngestResult:
        return IngestResult(document=parse_excel(path))


def _cell(row: tuple, index: int) -> object:
    """Safe cell access even for short/empty rows (None if absent)."""
    if not row or index >= len(row):
        return None
    return row[index]


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _none_if_blank(value: object) -> str | None:
    text = _clean(value)
    return text or None


def _guard_zip(path: str | Path) -> None:
    """Reject decompression-bomb workbooks from the ZIP central directory (no inflation).

    Sizes and ratios come from the central-directory headers, so this reads no member
    payload. A non-ZIP input (e.g. a legacy .xls) is left for openpyxl to reject.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
    except (zipfile.BadZipFile, OSError):
        return
    if len(infos) > _MAX_ZIP_ENTRIES:
        raise ParseError(f"Excel workbook has too many internal entries (> {_MAX_ZIP_ENTRIES})")
    total_uncompressed = sum(info.file_size for info in infos)
    total_compressed = sum(info.compress_size for info in infos)
    if total_uncompressed > _MAX_TOTAL_UNCOMPRESSED_BYTES:
        raise ParseError(
            "Excel workbook expands beyond the allowed size "
            f"({total_uncompressed:,} > {_MAX_TOTAL_UNCOMPRESSED_BYTES:,} bytes uncompressed)"
        )
    ratio = total_uncompressed / max(total_compressed, 1)
    if ratio > _MAX_COMPRESSION_RATIO:
        raise ParseError(
            f"Excel workbook compression ratio {ratio:.0f}x exceeds the allowed "
            f"{_MAX_COMPRESSION_RATIO:.0f}x (possible decompression bomb)"
        )


def parse_excel(path: str | Path) -> NoticeDocument:
    # A zip magic match (.zip/.docx/.jar) can reach this adapter, so openpyxl failures and
    # missing-sheet errors are wrapped as ParseError (400) instead of surfacing as HTTP 500.
    _guard_zip(path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises BadZipFile/InvalidFileException/etc.
        raise ParseError(f"failed to open Excel workbook: {Path(path).name}") from exc
    try:
        missing = [s for s in _REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            raise ParseError(f"Excel workbook is missing required sheet(s): {', '.join(missing)}")
        name = _document_name(wb)
        packages = _packages(wb)
        refs = _extracted_licenses(wb)
    finally:
        wb.close()
    return NoticeDocument(name=name, packages=tuple(packages), license_refs=tuple(refs))


def _document_name(wb: openpyxl.Workbook) -> str:
    ws = wb["Document Info"]
    rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if rows:
        return _clean(_cell(rows[0], _DOC_NAME_COL)) or "OSS Notice"
    return "OSS Notice"


def _packages(wb: openpyxl.Workbook) -> list[Package]:
    ws = wb["Package Info"]
    out: list[Package] = []
    # Deduplicate by (name, version) during the scan so a workbook full of repeated rows
    # never builds more than the distinct set of Package objects. NoticeDocument dedups by
    # the same key, but only after every object already exists; doing it here bounds peak
    # memory. The row counter aborts before the scan itself becomes the exhaustion vector.
    seen: set[tuple[str, str]] = set()
    scanned = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _clean(_cell(row, _PKG_COLS["name"]))
        if not name:
            continue
        scanned += 1
        if scanned > _MAX_PACKAGE_ROWS:
            raise ParseError(f"Package Info sheet has too many rows (> {_MAX_PACKAGE_ROWS})")
        version = _clean(_cell(row, _PKG_COLS["version"]))
        key = (name, version)
        if key in seen:
            continue
        seen.add(key)
        out.append(
            Package(
                name=name,
                version=version,
                license_concluded=LicenseExpression.from_raw(_cell(row, _PKG_COLS["concluded"])),
                license_declared=LicenseExpression.from_raw(_cell(row, _PKG_COLS["declared"])),
                copyright=Copyright.from_raw(_cell(row, _PKG_COLS["copyright"])),
                download_location=_clean(_cell(row, _PKG_COLS["download"])),
            )
        )
    return out


def _extracted_licenses(wb: openpyxl.Workbook) -> list[LicenseRef]:
    if "Extracted License Info" not in wb.sheetnames:
        return []
    ws = wb["Extracted License Info"]
    out: list[LicenseRef] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        identifier = _clean(_cell(row, _EXT_COLS["id"]))
        if not identifier:
            continue
        out.append(
            LicenseRef(
                identifier=identifier,
                name=_none_if_blank(_cell(row, _EXT_COLS["name"])),
                extracted_text=_clean(_cell(row, _EXT_COLS["text"])),
            )
        )
    return out
